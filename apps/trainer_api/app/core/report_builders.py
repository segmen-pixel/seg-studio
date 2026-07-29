# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Segmen-Pixel and Seg-Studio contributors
"""Building blocks for model-evaluation and batch-inspection reports.

Extracted verbatim from report_generator.py during the pre-OSS refactor:
training-log epoch parser, matplotlib chart builders (learning curves /
confusion matrix / threshold sweep / score distribution), metrics helpers
(model hash, instance recall, hard-case selection, per-class rows), image
embedding, Excel generation, HTML render + PDF conversion, project info,
and report label i18n.
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import logging
import re
from pathlib import Path
from typing import Any

import numpy as np

from ..core.paths import project_dir

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Epoch history parser
# ---------------------------------------------------------------------------
_EPOCH_RE = re.compile(
    r"Epoch\s+(\d+)/(\d+)\s+"
    r"- loss:\s+([\d.]+)\s+"
    r"- lr:\s+([\d.eE+-]+)\s+"
    r"- val mIoU:\s+([\d.]+)\s+"
    r"- val F1:\s+([\d.]+)"
    r"(?:\s+- distill:\s+([\d.]+))?"
    r"(?:\s+- opt_thresh:\s+([\d.]+)\s+\(F1=([\d.]+)\))?"
    r"(?:\s+- ECE:\s+([\d.]+))?"
)


def _parse_epoch_history(log_path: Path) -> list[dict[str, Any]]:
    """Parse train.log and extract per-epoch metrics."""
    if not log_path.exists():
        return []
    history: list[dict[str, Any]] = []
    text = log_path.read_text(encoding="utf-8", errors="replace")
    for m in _EPOCH_RE.finditer(text):
        entry: dict[str, Any] = {
            "epoch": int(m.group(1)),
            "total_epochs": int(m.group(2)),
            "loss": float(m.group(3)),
            "lr": float(m.group(4)),
            "val_miou": float(m.group(5)),
            "val_f1": float(m.group(6)),
        }
        if m.group(7) is not None:
            entry["distill_loss"] = float(m.group(7))
        if m.group(8) is not None:
            entry["opt_threshold"] = float(m.group(8))
            entry["opt_threshold_f1"] = float(m.group(9))
        if m.group(10) is not None:
            entry["ece"] = float(m.group(10))
        history.append(entry)
    return history


# ---------------------------------------------------------------------------
# Chart builders (matplotlib -> base64 PNG)
# ---------------------------------------------------------------------------
def _fig_to_base64(fig) -> str:
    """Convert matplotlib figure to base64-encoded PNG string."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode("ascii")
    import matplotlib.pyplot as plt
    plt.close(fig)
    return f"data:image/png;base64,{b64}"


def _build_learning_curves(history: list[dict]) -> str:
    """Build loss/F1/mIoU learning curves chart."""
    if not history:
        return ""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    epochs = [h["epoch"] for h in history]
    losses = [h["loss"] for h in history]
    f1s = [h["val_f1"] for h in history]
    mious = [h["val_miou"] for h in history]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

    # Loss curve
    ax1.plot(epochs, losses, "b-", linewidth=1.5, label="Loss")
    if "distill_loss" in history[0]:
        distill = [h.get("distill_loss", 0) for h in history]
        ax1.plot(epochs, distill, "r--", linewidth=1, alpha=0.7, label="Distill Loss")
    ax1.set_xlabel("Epoch")
    ax1.set_ylabel("Loss")
    ax1.set_title("Training Loss")
    ax1.legend()
    ax1.grid(True, alpha=0.3)

    # F1 & mIoU curves
    ax2.plot(epochs, f1s, "g-", linewidth=1.5, label="Val F1")
    ax2.plot(epochs, mious, "m-", linewidth=1.5, label="Val mIoU")
    if any("opt_threshold_f1" in h for h in history):
        opt_f1s = [h.get("opt_threshold_f1", h["val_f1"]) for h in history]
        ax2.plot(epochs, opt_f1s, "g--", linewidth=1, alpha=0.7, label="Opt-Thresh F1")
    ax2.set_xlabel("Epoch")
    ax2.set_ylabel("Score")
    ax2.set_title("Validation Metrics")
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(0, 1.05)

    fig.tight_layout()
    return _fig_to_base64(fig)


def _build_confusion_matrix_chart(
    cm: list[list[float]], class_names: list[str]
) -> str:
    """Build confusion matrix heatmap chart."""
    if not cm:
        return ""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    arr = np.array(cm, dtype=np.float64)
    # Normalize per row (true label)
    row_sums = arr.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1
    arr_norm = arr / row_sums

    fig, ax = plt.subplots(figsize=(max(5, len(class_names) * 1.2), max(4, len(class_names))))
    im = ax.imshow(arr_norm, cmap="Blues", vmin=0, vmax=1, aspect="auto")
    fig.colorbar(im, ax=ax, shrink=0.8)

    ax.set_xticks(range(len(class_names)))
    ax.set_yticks(range(len(class_names)))
    ax.set_xticklabels(class_names, rotation=45, ha="right", fontsize=8)
    ax.set_yticklabels(class_names, fontsize=8)
    ax.set_xlabel("Predicted")
    ax.set_ylabel("True")
    ax.set_title("Confusion Matrix (normalized)")

    # Annotate cells
    for i in range(len(class_names)):
        for j in range(len(class_names)):
            val = arr_norm[i, j]
            count = int(arr[i, j])
            color = "white" if val > 0.5 else "black"
            ax.text(j, i, f"{val:.2f}\n({count})", ha="center", va="center",
                    fontsize=7, color=color)

    fig.tight_layout()
    return _fig_to_base64(fig)


def _build_threshold_chart(
    history: list[dict],
) -> str:
    """Build threshold vs F1 chart from epoch history with threshold data."""
    entries = [h for h in history if "opt_threshold" in h and "opt_threshold_f1" in h]
    if not entries:
        return ""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    epochs = [e["epoch"] for e in entries]
    thresholds = [e["opt_threshold"] for e in entries]
    eces = [e.get("ece", 0) for e in entries]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

    ax1.plot(epochs, thresholds, "b-o", markersize=3, linewidth=1.5)
    ax1.set_xlabel("Epoch")
    ax1.set_ylabel("Optimal Threshold")
    ax1.set_title("Optimal Threshold over Training")
    ax1.set_ylim(0, 1)
    ax1.grid(True, alpha=0.3)

    ax2.plot(epochs, eces, "r-o", markersize=3, linewidth=1.5)
    ax2.set_xlabel("Epoch")
    ax2.set_ylabel("ECE")
    ax2.set_title("Expected Calibration Error")
    ax2.grid(True, alpha=0.3)

    fig.tight_layout()
    return _fig_to_base64(fig)


# ---------------------------------------------------------------------------
# Model hash (traceability)
# ---------------------------------------------------------------------------
def _compute_model_hash(model_path: Path) -> str:
    """Compute SHA256 hash of model checkpoint."""
    if not model_path.exists():
        return "N/A"
    sha = hashlib.sha256()
    with open(model_path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            sha.update(chunk)
    return sha.hexdigest()


# ---------------------------------------------------------------------------
# Instance-level recall (connected component based)
# ---------------------------------------------------------------------------
def _compute_instance_recall(
    gt_masks_dir: Path,
    pred_masks_dir: Path,
    class_ids: list[int],
    threshold: float = 0.5,
) -> dict[str, Any]:
    """Compute instance-level recall using connected components.

    For each defect instance (connected component) in GT, check if
    the prediction overlaps with IoU >= threshold.
    """
    from PIL import Image
    from scipy.ndimage import label as ndimage_label

    total_instances = 0
    detected_instances = 0
    missed: list[dict] = []

    # Background is not a defect instance. Callers pass the project's ACTIVE
    # class ids, and class 0 "background" is active in every project, so it
    # arrived here as one enormous connected component that the model also
    # predicts as background -- a guaranteed +1 detected and +1 total on every
    # image. With a handful of real defects that alone lifts the reported recall
    # a long way: two real instances with one missed reads 0.67 instead of 0.50.
    # Every other metric in the codebase is foreground-only (include_background
    # =False); this one is too, and it enforces that here rather than trusting
    # each caller.
    fg_class_ids = [int(c) for c in class_ids if int(c) != 0]

    gt_files = sorted(gt_masks_dir.glob("*.png"))
    for gt_path in gt_files:
        stem = gt_path.stem
        pred_path = pred_masks_dir / f"{stem}.png"
        if not pred_path.exists():
            continue

        gt_mask = np.array(Image.open(gt_path))
        pred_mask = np.array(Image.open(pred_path))

        for cid in fg_class_ids:
            gt_binary = (gt_mask == cid).astype(np.uint8)
            pred_binary = (pred_mask == cid).astype(np.uint8)
            if gt_binary.sum() == 0:
                continue

            labeled, n_instances = ndimage_label(gt_binary)
            for inst_id in range(1, n_instances + 1):
                inst_mask = (labeled == inst_id)
                total_instances += 1
                # Check overlap
                intersection = (inst_mask & (pred_binary > 0)).sum()
                # Denominator is THIS instance's area, i.e. how much of the
                # instance the model covered. It must not include predictions
                # outside the instance, or every other defect predicted in the
                # same image inflates it and pushes real detections into
                # "missed".
                #
                # This read `pred_binary > 0 & inst_mask`, which Python parses
                # as `pred_binary > (0 & inst_mask)` -- `&` binds tighter than
                # `>` -- and `0 & inst_mask` is all-False, so the whole term
                # collapsed to `pred_binary > 0` and the union became the
                # instance plus the model's ENTIRE predicted mask for the
                # class across the image, so an instance's score fell as the
                # model found MORE defects. Three equal instances, all predicted
                # perfectly, scored 1/3 each and were all three reported missed;
                # two scraped by only because 0.5 passes the >= 0.5 threshold
                # exactly.
                union = int(inst_mask.sum())
                iou = intersection / max(union, 1)
                if iou >= threshold:
                    detected_instances += 1
                else:
                    missed.append({
                        "image": stem,
                        "class_id": cid,
                        "instance_area_px": int(inst_mask.sum()),
                        "iou": float(iou),
                    })

    recall = detected_instances / max(total_instances, 1)
    return {
        "total_instances": total_instances,
        "detected_instances": detected_instances,
        "instance_recall": round(recall, 4),
        "missed_instances": sorted(missed, key=lambda x: x["instance_area_px"], reverse=True),
    }


# ---------------------------------------------------------------------------
# Hard case selection
# ---------------------------------------------------------------------------
def _select_hard_cases(
    preds_dir: Path,
    top_n: int = 10,
) -> dict[str, list[dict]]:
    """Select hard cases: lowest confidence, highest area FP/FN."""
    scores: list[dict] = []
    for score_path in preds_dir.glob("*.score.json"):
        try:
            data = json.loads(score_path.read_text(encoding="utf-8"))
            data["_stem"] = score_path.stem.replace(".score", "")
            scores.append(data)
        except Exception:
            continue

    if not scores:
        return {"low_confidence": [], "high_fg_ratio": []}

    # Sort by mean_confidence ascending (low confidence = hard)
    by_confidence = sorted(scores, key=lambda x: x.get("mean_confidence", 1.0))
    low_confidence = by_confidence[:top_n]

    # Sort by foreground_ratio descending (high FP area)
    by_fg = sorted(scores, key=lambda x: x.get("foreground_ratio", 0), reverse=True)
    high_fg = by_fg[:top_n]

    return {
        "low_confidence": [
            {"image": s["_stem"], "mean_confidence": s.get("mean_confidence", 0),
             "fg_ratio": s.get("foreground_ratio", 0)}
            for s in low_confidence
        ],
        "high_fg_ratio": [
            {"image": s["_stem"], "fg_ratio": s.get("foreground_ratio", 0),
             "mean_confidence": s.get("mean_confidence", 0)}
            for s in high_fg
        ],
    }


# ---------------------------------------------------------------------------
# Image encoding for hard case display
# ---------------------------------------------------------------------------
def _encode_image_base64(image_path: Path, max_size: int = 512) -> str:
    """Read image, resize to max_size, return base64 PNG."""
    if not image_path.exists():
        return ""
    from PIL import Image
    img = Image.open(image_path)
    img.thumbnail((max_size, max_size), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return f"data:image/png;base64,{base64.b64encode(buf.read()).decode('ascii')}"


def _build_hard_case_images(
    hard_cases: dict,
    images_dir: Path,
    preds_dir: Path,
    max_items: int = 5,
) -> list[dict]:
    """Build 3-panel images for hard cases (original, prediction overlay, confidence map)."""
    items: list[dict] = []
    seen = set()
    for case in hard_cases.get("low_confidence", [])[:max_items]:
        stem = case["image"]
        if stem in seen:
            continue
        seen.add(stem)

        # Find original image
        orig_path = None
        for ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff"):
            candidate = images_dir / f"{stem}{ext}"
            if candidate.exists():
                orig_path = candidate
                break

        pred_path = preds_dir / f"{stem}.png"
        conf_path = preds_dir / f"{stem}.confidence.png"

        items.append({
            "stem": stem,
            "mean_confidence": case.get("mean_confidence", 0),
            "fg_ratio": case.get("fg_ratio", 0),
            "original": _encode_image_base64(orig_path) if orig_path else "",
            "prediction": _encode_image_base64(pred_path),
            "confidence": _encode_image_base64(conf_path),
        })
    return items


# ---------------------------------------------------------------------------
# Score distribution chart
# ---------------------------------------------------------------------------
def _build_score_distribution(preds_dir: Path) -> str:
    """Build confidence score distribution histogram."""
    confidences: list[float] = []
    for score_path in preds_dir.glob("*.score.json"):
        try:
            data = json.loads(score_path.read_text(encoding="utf-8"))
            confidences.append(data.get("mean_confidence", 0))
        except Exception:
            continue
    if not confidences:
        return ""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.hist(confidences, bins=20, color="#4a90d9", edgecolor="white", alpha=0.8)
    ax.set_xlabel("Mean Confidence")
    ax.set_ylabel("Count")
    ax.set_title("Prediction Confidence Distribution")
    ax.axvline(np.mean(confidences), color="red", linestyle="--", label=f"Mean: {np.mean(confidences):.3f}")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return _fig_to_base64(fig)


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------
def _generate_excel(
    report_data: dict[str, Any],
    output_path: Path,
) -> Path:
    """Generate Excel report with summary + per-class stats + hard cases."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws["A1"] = report_data.get("title", "Model Evaluation Report")
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    # KPIs (accept either a dict or a list of {label, value})
    kpis = report_data.get("kpis", {})
    kpi_items = kpis.items() if isinstance(kpis, dict) else [(k["label"], k["value"]) for k in kpis]
    row = 3
    for key, val in kpi_items:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(val))
        row += 1

    # --- Per-class sheet ---
    ws2 = wb.create_sheet("Per-Class Metrics")
    class_headers = ["Class", "Precision", "Recall", "F1", "IoU", "Support (px)"]
    for col, h in enumerate(class_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    per_class = report_data.get("per_class_metrics", [])
    for i, cls in enumerate(per_class, 2):
        ws2.cell(row=i, column=1, value=cls.get("name", "")).border = thin_border
        ws2.cell(row=i, column=2, value=cls.get("precision", 0)).border = thin_border
        ws2.cell(row=i, column=3, value=cls.get("recall", 0)).border = thin_border
        ws2.cell(row=i, column=4, value=cls.get("f1", 0)).border = thin_border
        ws2.cell(row=i, column=5, value=cls.get("iou", 0)).border = thin_border
        ws2.cell(row=i, column=6, value=cls.get("support_px", 0)).border = thin_border
        # Number format
        for col in range(2, 6):
            ws2.cell(row=i, column=col).number_format = "0.0000"

    # Auto-width
    for col_cells in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = max(max_len + 2, 12)

    # --- Hard cases sheet ---
    if report_data.get("hard_cases"):
        ws3 = wb.create_sheet("Hard Cases")
        hc_headers = ["Image", "Mean Confidence", "FG Ratio"]
        for col, h in enumerate(hc_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for i, case in enumerate(report_data["hard_cases"].get("low_confidence", []), 2):
            ws3.cell(row=i, column=1, value=case.get("image", "")).border = thin_border
            ws3.cell(row=i, column=2, value=case.get("mean_confidence", 0)).border = thin_border
            ws3.cell(row=i, column=3, value=case.get("fg_ratio", 0)).border = thin_border

    # --- Training history sheet ---
    history = report_data.get("epoch_history", [])
    if history:
        ws4 = wb.create_sheet("Training History")
        hist_headers = ["Epoch", "Loss", "LR", "Val mIoU", "Val F1"]
        for col, h in enumerate(hist_headers, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for i, h in enumerate(history, 2):
            ws4.cell(row=i, column=1, value=h["epoch"]).border = thin_border
            ws4.cell(row=i, column=2, value=h["loss"]).border = thin_border
            ws4.cell(row=i, column=3, value=h["lr"]).border = thin_border
            ws4.cell(row=i, column=4, value=h["val_miou"]).border = thin_border
            ws4.cell(row=i, column=5, value=h["val_f1"]).border = thin_border
            ws4.cell(row=i, column=2).number_format = "0.0000"
            ws4.cell(row=i, column=3).number_format = "0.00E+00"
            ws4.cell(row=i, column=4).number_format = "0.0000"
            ws4.cell(row=i, column=5).number_format = "0.0000"

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# HTML rendering (Jinja2)
# ---------------------------------------------------------------------------
def _render_html(template_name: str, context: dict[str, Any]) -> str:
    """Render HTML report from Jinja2 template."""
    from jinja2 import Environment, FileSystemLoader

    templates_dir = Path(__file__).parent.parent / "templates"
    env = Environment(
        loader=FileSystemLoader(str(templates_dir)),
        autoescape=True,
    )
    template = env.get_template(template_name)
    return template.render(**context)


# ---------------------------------------------------------------------------
# PDF generation (WeasyPrint)
# ---------------------------------------------------------------------------
def _html_to_pdf(html_content: str, output_path: Path) -> Path | None:
    """Convert HTML string to PDF using WeasyPrint."""
    try:
        from weasyprint import HTML
        HTML(string=html_content).write_pdf(str(output_path))
        return output_path
    except ImportError:
        logger.warning("WeasyPrint not installed — skipping PDF generation")
        return None
    except Exception as exc:
        logger.warning("PDF generation failed: %s", exc)
        return None


# ---------------------------------------------------------------------------
# Main report generators
# ---------------------------------------------------------------------------
def _load_project_info(project_id: str) -> dict[str, Any]:
    """Load project.json and classes.json."""
    pdir = project_dir(project_id)
    info: dict[str, Any] = {"project_id": project_id}

    proj_path = pdir / "project.json"
    if proj_path.exists():
        proj = json.loads(proj_path.read_text(encoding="utf-8"))
        info["project_name"] = proj.get("name", project_id)
        info["created_at"] = proj.get("created_at", "")
    else:
        info["project_name"] = project_id

    classes_path = pdir / "classes.json"
    if classes_path.exists():
        classes_data = json.loads(classes_path.read_text(encoding="utf-8"))
        classes_list = classes_data.get("classes", [])
        info["classes"] = classes_list
        # Build id->name map
        info["class_map"] = {str(c["id"]): c["name"] for c in classes_list}
    else:
        info["classes"] = []
        info["class_map"] = {}

    return info


def _build_per_class_metrics(
    metrics: dict, class_map: dict[str, str]
) -> list[dict]:
    """Build per-class metrics table from metrics.json data."""
    per_class_f1 = metrics.get("per_class_f1_val", {})
    per_class_prec = metrics.get("per_class_precision_val", {})
    per_class_rec = metrics.get("per_class_recall_val", {})
    per_class_iou = metrics.get("per_class_iou_val", {})

    rows: list[dict] = []
    for cid in sorted(per_class_f1.keys(), key=lambda x: int(x)):
        rows.append({
            "id": cid,
            "name": class_map.get(cid, f"class_{cid}"),
            "precision": round(float(per_class_prec.get(cid, 0)), 4),
            "recall": round(float(per_class_rec.get(cid, 0)), 4),
            "f1": round(float(per_class_f1.get(cid, 0)), 4),
            "iou": round(float(per_class_iou.get(cid, 0)), 4),
        })
    return rows


def _report_labels(lang: str) -> dict[str, str]:
    """Localized UI strings for the HTML report (ja / en)."""
    ja = lang == "ja"
    return {
        "report_title": "モデル評価レポート" if ja else "Model Evaluation Report",
        "run": "ラン" if ja else "Run",
        "generated": "生成日時" if ja else "Generated",
        "summary": "サマリー" if ja else "Summary",
        "summary_intro": "学習結果の主要指標です。F1 は適合率と再現率のバランス（1.0 が最良）。" if ja else "Key training metrics. F1 balances precision and recall (1.0 is best).",
        "kpi_best_f1": "ベスト F1" if ja else "Best F1",
        "kpi_best_miou": "ベスト mIoU" if ja else "Best mIoU",
        "kpi_best_epoch": "ベストエポック" if ja else "Best Epoch",
        "kpi_total_epochs": "総エポック数" if ja else "Total Epochs",
        "kpi_final_loss": "最終 Loss" if ja else "Final Loss",
        "kpi_optimal_threshold": "最適しきい値" if ja else "Optimal Threshold",
        "kpi_ece": "ECE（較正誤差）" if ja else "ECE",
        "kpi_train_images": "学習画像数" if ja else "Train Images",
        "kpi_val_images": "検証画像数" if ja else "Val Images",
        "kpi_input_size": "入力サイズ" if ja else "Input Size",
        "kpi_architecture": "アーキテクチャ" if ja else "Architecture",
        "judg_good_strong": "デプロイ推奨。" if ja else "Recommended for deployment.",
        "judg_good_body": "が 0.90 を超えています。" if ja else "exceeds the 0.90 threshold.",
        "judg_warn_strong": "追加学習を検討。" if ja else "Consider additional training.",
        "judg_warn_body": "は中程度です。下の Hard Case を確認してください。" if ja else "is moderate. Review the hard cases below.",
        "judg_bad_strong": "デプロイ非推奨。" if ja else "Not recommended for deployment.",
        "judg_bad_body": "は許容しきい値を下回ります。しきい値の再調整かデータ追加を推奨します。" if ja else "is below the acceptable threshold. Re-calibrate the threshold or add more data.",
        "per_class": "クラス別性能" if ja else "Per-Class Performance",
        "per_class_intro": "各クラスの適合率・再現率・F1・IoU。" if ja else "Precision, recall, F1 and IoU per class.",
        "col_class": "クラス" if ja else "Class",
        "col_precision": "適合率" if ja else "Precision",
        "col_recall": "再現率" if ja else "Recall",
        "col_f1": "F1",
        "col_iou": "IoU",
        "no_per_class": "クラス別指標がありません。" if ja else "No per-class metrics available.",
        "instance_recall": "インスタンス単位の再現率" if ja else "Instance-level Recall",
        "instance_recall_intro": "欠陥を1個ずつ数えた検出率（IoU 0.5 以上で検出とみなす）。" if ja else "Detection rate counted per defect instance (detected when IoU >= 0.5).",
        "defects_detected": "個の欠陥を検出" if ja else "defect instances detected",
        "missed_title": "見逃したインスタンス（上位10件）" if ja else "Missed Instances (top 10)",
        "col_image": "画像" if ja else "Image",
        "col_area": "面積 (px)" if ja else "Area (px)",
        "training_progress": "学習推移" if ja else "Training Progress",
        "training_progress_intro": "エポックごとの Loss と検証指標の推移。" if ja else "Loss and validation metrics across epochs.",
        "confusion_matrix": "混同行列" if ja else "Confusion Matrix",
        "confusion_matrix_intro": "予測と正解の対応（対角＝正解、行で正規化）。" if ja else "Predicted vs. true classes (diagonal = correct, row-normalized).",
        "threshold_cal": "しきい値・較正" if ja else "Threshold & Calibration",
        "threshold_cal_intro": "最適しきい値の推移と較正誤差(ECE)。ECE は小さいほど確信度が正確です。" if ja else "Optimal threshold over training and calibration error (ECE; lower is better).",
        "confidence_dist": "確信度の分布" if ja else "Confidence Distribution",
        "confidence_dist_intro": "予測の確信度ヒストグラム。" if ja else "Histogram of prediction confidence.",
        "hard_cases": "Hard Case（低確信度）" if ja else "Hard Cases (Low Confidence)",
        "hard_cases_intro": "確信度が低く誤りやすい画像。元画像／予測／確信度マップを並べています。" if ja else "Low-confidence, error-prone images: original / prediction / confidence map.",
        "confidence": "確信度" if ja else "Confidence",
        "fg": "前景率" if ja else "FG",
        "panel_original": "元画像" if ja else "Original",
        "panel_prediction": "予測" if ja else "Prediction",
        "panel_confidence": "確信度マップ" if ja else "Confidence Map",
        "dataset_health": "データセット概況" if ja else "Dataset Health",
        "ds_train": "学習画像数" if ja else "Training Images",
        "ds_val": "検証画像数" if ja else "Validation Images",
        "ds_classes": "クラス数" if ja else "Number of Classes",
        "ds_mean_size": "平均画像サイズ" if ja else "Mean Image Size",
        "ds_fg": "前景率" if ja else "Foreground Ratio",
        "ds_input": "入力サイズ" if ja else "Input Size",
        "ds_stride": "出力ストライド" if ja else "Output Stride",
        "traceability": "トレーサビリティ" if ja else "Traceability",
        "tr_model_hash": "モデルハッシュ (SHA-256)" if ja else "Model Hash (SHA-256)",
        "tr_run_id": "ラン ID" if ja else "Run ID",
        "tr_project_id": "プロジェクト ID" if ja else "Project ID",
        "tr_arch": "アーキテクチャ" if ja else "Architecture",
        "tr_loss": "損失関数" if ja else "Loss Type",
        "tr_lr": "学習率" if ja else "Learning Rate",
        "tr_batch": "バッチサイズ" if ja else "Batch Size",
        "tr_patch": "パッチサイズ" if ja else "Patch Size",
        "tr_epochs": "エポック数" if ja else "Epochs",
        "tr_distill": "蒸留" if ja else "Distillation",
        "tr_aug": "データ拡張" if ja else "Augmentation",
        "tr_generated": "レポート生成日時" if ja else "Report Generated",
        "tr_version": "Seg-Studio バージョン" if ja else "Seg-Studio Version",
        "enabled": "有効" if ja else "Enabled",
        "disabled": "無効" if ja else "Disabled",
        "footer_by": "生成: Seg-Studio" if ja else "Generated by Seg-Studio",
    }
